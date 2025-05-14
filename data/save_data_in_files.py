import json
import os.path

import yaml
from openpyxl import load_workbook, Workbook

class SaveData:

    @staticmethod
    def save_json(user_file=None, payload=None):
        users = []
        if os.path.exists(user_file):
            with open(user_file, "r") as f:
                users = json.load(f)
        users.append(payload)
        with open(user_file, "w") as f:
            json.dump(users, f, indent=4)

    @staticmethod
    def save_excel(user_file=None, payload=None):
        if os.path.exists(user_file):
            workbook=load_workbook(user_file)
            sheet=workbook.active

        else:
            workbook=Workbook()
            sheet=workbook.active
            sheet.append(list(payload.keys()))
        sheet.append(list(payload.values()))
        workbook.save(user_file)

    @staticmethod
    def save_yaml(user_file,payload):
        users=[]
        if os.path.exists(user_file):
            with open(user_file, "r") as f :
                users=yaml.safe_load(f)
        users.append(payload)
        with open(user_file, "w") as f:
            yaml.safe_dump(users,f)

